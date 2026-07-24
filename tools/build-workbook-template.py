# -*- coding: utf-8 -*-
"""Build workbook-template.json from the official "PNPT Configuration
Workbook _ NAMER.xlsx".

The template captures the ENTIRE workbook 1:1 — every tab, every row with its
kind (title / header / black tool banner / gray sub-group / data), per-tab
column widths, row heights, merge flags, and the PROCORE logo PNG + drawing
anchor. It is the single source of truth for BOTH:
  - the Config Tracker's Build-phase workbook UI (config-tracker.js derives
    the active tier's sections/rows from it), and
  - the .xlsx / Google Sheets export (workbook-export.js rebuilds the file
    from it and injects the tracked values by exact row reference).

Regenerate whenever the official workbook changes:
  python tools/build-workbook-template.py

Reads:  ~/Documents/PNPT TRAINING DOCS/PNPT Configuration Workbook _ NAMER.xlsx
Writes: workbook-template.json (committed; loaded lazily by the app)
"""
import base64, json, os, sys, zipfile
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(os.path.expanduser("~"), "Documents", "PNPT TRAINING DOCS",
                    "PNPT Configuration Workbook _ NAMER.xlsx")
OUT = os.path.join(ROOT, "workbook-template.json")

# Config Tracker tier key -> workbook tab name
TIER_TABS = {
    "standard": "Cost Management",
    "enterprise": "Cost Management Enterprise",
    "pe-essentials": "Project Execution Essentials",
    "pe-standard": "Project Execution",
    "pe-enterprise": "Project Execution Enterprise",
    "rm-tracking": "Resource Tracking",
    "rm-planning": "Resource Planning",
    "rm-advanced": "Resource Management",
    "plm-owners": "Project Lifecycle Management",
}

def rgb(color):
    try:
        v = color.rgb
        return str(v) if v is not None else None
    except Exception:
        return None

print("Reading", XLSX)
wb = openpyxl.load_workbook(XLSX)
z = zipfile.ZipFile(XLSX)

logo_b64 = base64.b64encode(z.read("xl/media/image1.png")).decode("ascii")
drawing_xml = z.read("xl/drawings/drawing1.xml").decode("utf-8")

tabs = []
for ws in wb.worksheets:
    merged = set()
    for m in ws.merged_cells.ranges:
        if m.min_col == 1 and m.max_col >= 5:
            merged.add(m.min_row)

    # ------- resolve this tab's column layout by HEADER TEXT (row 3) -------
    # Tabs are NOT uniform: PE + Resource tabs are 5-col
    # (Updated=C, Changed to=D, Notes=E); Cost Management, CM Enterprise, and
    # PLM are 6-col — an extra "Discussion Point / Decision Logic" column B
    # shifts Updated=D, Changed to=E, Notes=F. Everything downstream keys off
    # these resolved roles, never fixed letters.
    header_row = None
    for r in range(1, 8):
        if str(ws.cell(row=r, column=1).value or "").strip().lower() == "discussion point":
            header_row = r
            break
    roles = {}          # role -> 0-based column index
    if header_row is not None:
        for c in range(1, 12):
            h = str(ws.cell(row=header_row, column=c).value or "").strip().lower()
            if h == "discussion point" and "discussion" not in roles:
                roles["discussion"] = c - 1
            elif h == "discussion point / decision logic":
                roles["decisionLogic"] = c - 1
            elif h == "default setting":
                roles["default"] = c - 1
            elif h == "updated":
                roles["updated"] = c - 1
            elif h == "changed to":
                roles["changedTo"] = c - 1
            elif h == "notes":
                roles["notes"] = c - 1
        for req in ("discussion", "default", "updated", "changedTo", "notes"):
            if req not in roles:
                sys.exit("ABORT: tab %r missing header role %r (header row %d)" % (ws.title, req, header_row))
        ncols = max(roles.values()) + 1
    else:
        # Non-tier placeholder tabs (Correspondence/Inspection/Action Plan/
        # Workflows templates) — no standard header; keep verbatim, no roles.
        ncols = 0
        for r in range(1, ws.max_row + 1):
            for c in range(1, 12):
                if ws.cell(row=r, column=c).value not in (None, ""):
                    ncols = max(ncols, c)
        ncols = max(ncols, 1)

    last = 3
    for r in range(4, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, ncols + 1)):
            last = r

    widths = {}
    for i in range(ncols):
        L = chr(65 + i)
        d = ws.column_dimensions.get(L)
        if d is not None and d.width:
            widths[L] = round(float(d.width), 2)

    up_idx = roles.get("updated", -1)
    rows = []
    for r in range(1, last + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
        a_cell = ws.cell(row=r, column=1)
        fill = a_cell.fill
        fg = rgb(fill.fgColor) if (fill is not None and fill.patternType == "solid") else None
        if r <= 2:
            kind = "title"
        elif header_row is not None and r == header_row:
            kind = "header"
        elif fg == "FF000000":
            kind = "banner"
        elif fg == "FFE5E5E5":
            kind = "sub"
        else:
            kind = "data"
        hdim = ws.row_dimensions.get(r)
        h = round(float(hdim.height), 2) if (hdim is not None and hdim.height) else None
        # Normalize the Updated (checkbox) cell on data rows: boolean where the
        # sheet has TRUE/FALSE, verbatim otherwise ("N/A").
        v = []
        for i, cv in enumerate(cells):
            if kind == "data" and i == up_idx:
                if isinstance(cv, bool):
                    v.append(cv)
                elif isinstance(cv, str) and cv.strip().upper() in ("TRUE", "FALSE"):
                    v.append(cv.strip().upper() == "TRUE")
                else:
                    v.append(cv)  # None or verbatim text
            else:
                v.append(cv if cv is not None else "")
        row = {"r": r, "k": kind, "v": v}
        if r in merged:
            row["m"] = True
        if h is not None:
            row["h"] = h
        rows.append(row)

    tabs.append({
        "name": ws.title,
        "freeze": bool(ws.freeze_panes),
        "ncols": ncols,
        "roles": roles,          # role -> 0-based column index
        "widths": widths,
        "rows": rows,
    })
    print("  tab %-32s rows=%-4d  (banners=%d subs=%d data=%d)" % (
        ws.title, last,
        sum(1 for x in rows if x["k"] == "banner"),
        sum(1 for x in rows if x["k"] == "sub"),
        sum(1 for x in rows if x["k"] == "data")))

template = {
    "_generated": "by tools/build-workbook-template.py from 'PNPT Configuration Workbook _ NAMER.xlsx'",
    "tierTabs": TIER_TABS,
    "logoPng": logo_b64,
    "drawingXml": drawing_xml,
    "tabs": tabs,
}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(template, f, ensure_ascii=False, separators=(",", ":"))
print("Wrote %s (%.0f KB)" % (OUT, os.path.getsize(OUT) / 1024.0))
